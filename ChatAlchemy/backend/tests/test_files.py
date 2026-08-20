from io import BytesIO

from openpyxl import Workbook, load_workbook

from chatalchemy.files import MAX_ROWS, export_xlsx_bytes, extract_document_text, parse_tabular_bytes


def test_parse_csv_bytes():
    rows = parse_tabular_bytes("sample.csv", b"drug,status\nosimertinib,approved\ngefitinib,approved\n")
    assert rows == [
        {"drug": "osimertinib", "status": "approved"},
        {"drug": "gefitinib", "status": "approved"},
    ]


def test_csv_parsing_is_bounded_to_max_rows():
    payload = "value\n" + "\n".join(str(index) for index in range(MAX_ROWS + 50)) + "\n"
    rows = parse_tabular_bytes("large.csv", payload.encode())
    assert len(rows) == MAX_ROWS
    assert rows[0]["value"] == "0"
    assert rows[-1]["value"] == str(MAX_ROWS - 1)


def test_parse_xlsx_all_sheets():
    wb = Workbook()
    ws = wb.active
    ws.title = "SheetA"
    ws.append(["gene", "score"])
    ws.append(["EGFR", 0.9])
    ws2 = wb.create_sheet("SheetB")
    ws2.append(["gene", "score"])
    ws2.append(["ALK", 0.8])
    buffer = BytesIO()
    wb.save(buffer)

    rows = parse_tabular_bytes("genes.xlsx", buffer.getvalue())
    assert rows[0]["gene"] == "EGFR"
    assert rows[0]["__sheet"] == "SheetA"
    assert rows[1]["gene"] == "ALK"
    assert rows[1]["__sheet"] == "SheetB"


def test_export_xlsx_bytes_roundtrip():
    payload = export_xlsx_bytes(["drug", "count"], [["pembrolizumab", 3]])
    wb = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    wb.close()
    assert values[0] == ("drug", "count")
    assert values[1] == ("pembrolizumab", 3)


def test_export_xlsx_treats_formula_like_text_as_literal_data():
    payload = export_xlsx_bytes(
        ["value"],
        [["=HYPERLINK(\"https://example.invalid\",\"click\")"], ["@SUM(1,2)"]],
    )
    wb = load_workbook(BytesIO(payload), read_only=True, data_only=False)
    values = list(wb.active.iter_rows(values_only=True))
    wb.close()
    assert values[1][0].startswith("'=")
    assert values[2][0].startswith("'@")


def test_extract_txt_document():
    assert extract_document_text("paper.txt", b"EGFR is discussed in lung cancer.") == "EGFR is discussed in lung cancer."
