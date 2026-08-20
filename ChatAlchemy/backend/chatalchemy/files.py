from __future__ import annotations

import csv
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any, Iterable

import xlrd
from openpyxl import Workbook, load_workbook
from pypdf import PdfReader

MAX_ROWS = 20_000
MAX_PDF_PAGES = 150
FORMULA_PREFIXES = ("=", "+", "-", "@")


def _json_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _xlsx_cell(value: Any) -> Any:
    """Write user/source text as data, not as an executable spreadsheet formula."""
    value = _json_cell(value)
    if isinstance(value, str):
        stripped = value.lstrip()
        if stripped.startswith(FORMULA_PREFIXES):
            return "'" + value
    return value


def _unique_headers(values: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(values):
        base = str(value).strip() if value not in (None, "") else f"column_{index + 1}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _rows_from_iterable(
    iterable: Iterable[list[Any] | tuple[Any, ...]],
    sheet: str | None = None,
    max_rows: int = MAX_ROWS,
) -> list[dict[str, Any]]:
    """Parse at most `max_rows` data rows without materializing a whole sheet."""
    headers: list[str] | None = None
    rows: list[dict[str, Any]] = []
    for raw_values in iterable:
        raw = list(raw_values)
        if headers is None:
            if not any(value not in (None, "") for value in raw):
                continue
            headers = _unique_headers(raw)
            continue

        if not any(value not in (None, "") for value in raw):
            continue
        padded = raw + [""] * max(0, len(headers) - len(raw))
        item = {headers[index]: _json_cell(padded[index]) for index in range(len(headers))}
        if sheet:
            item["__sheet"] = sheet
        rows.append(item)
        if len(rows) >= max_rows:
            break
    return rows


def _rows_from_matrix(matrix: list[list[Any]], sheet: str | None = None) -> list[dict[str, Any]]:
    return _rows_from_iterable(matrix, sheet=sheet, max_rows=MAX_ROWS)


def parse_tabular_bytes(filename: str, content: bytes) -> list[dict[str, Any]]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "csv":
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(StringIO(text))
        return _rows_from_iterable(reader, max_rows=MAX_ROWS)

    if ext == "xlsx":
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        out: list[dict[str, Any]] = []
        try:
            for ws in wb.worksheets:
                remaining = MAX_ROWS - len(out)
                if remaining <= 0:
                    break
                out.extend(
                    _rows_from_iterable(
                        ws.iter_rows(values_only=True),
                        sheet=ws.title,
                        max_rows=remaining,
                    )
                )
        finally:
            wb.close()
        return out

    if ext == "xls":
        wb = xlrd.open_workbook(file_contents=content)
        out: list[dict[str, Any]] = []
        for ws in wb.sheets():
            remaining = MAX_ROWS - len(out)
            if remaining <= 0:
                break
            out.extend(
                _rows_from_iterable(
                    (ws.row_values(index) for index in range(ws.nrows)),
                    sheet=ws.name,
                    max_rows=remaining,
                )
            )
        return out

    raise ValueError("Only CSV and Excel files (.csv, .xlsx, .xls) are supported")


def export_xlsx_bytes(headers: list[str], rows: list[list[Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append([_xlsx_cell(value) for value in headers])
    for row in rows[:MAX_ROWS]:
        ws.append([_xlsx_cell(value) for value in row])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def extract_document_text(filename: str, content: bytes, max_pages: int = MAX_PDF_PAGES) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "pdf":
        reader = PdfReader(BytesIO(content))
        pieces: list[str] = []
        for page in reader.pages[:max_pages]:
            text = page.extract_text() or ""
            if text.strip():
                pieces.append(text)
        return "\n".join(pieces).strip()
    if ext == "txt":
        return content.decode("utf-8-sig", errors="replace").strip()
    raise ValueError("Biomedical Analysis accepts PDF or TXT documents")
